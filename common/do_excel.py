# -*- coding: utf-8 -*-
# @time     : 2018/12/14 0014 下午 22:04
# @Author   : yuxuan
# #file     : do_excel.py

# excel表中每一条数据就是一个case，写一个case类，属性是id、title、url等等

import openpyxl
from common.http_requests_api import HttpRequest
import json
from common import contants
from common.read_config import ReadConfig

# Case类初始化属性,当excel列比较多，可采取如下方式，初始化属性写出来，后面读取excel再赋值上去
class Case:

    def __init__(self):
        self.case_id = None
        self.title = None
        self.url = None
        self.data = None
        self.method = None
        self.expected = None
        self.actual = None
        self.result = None


class DoExcel:

    #初始化函数可以把打开excel放进去
    def __init__(self,file_name):
        try:     #判断文件是否存在
            self.file_name = file_name
            self.wb = openpyxl.load_workbook(filename=file_name)
        except FileNotFoundError as e:
            print('{0} not found,please check file path'.format(file_name))
            raise e

    def get_case(self,sheet_name):
        sheet = self.wb[sheet_name]
        max_row = sheet.max_row
        cases = []
        for i in range(2, max_row + 1):
            case = Case()    #Case类实例化，用来存放测试数据
            case.case_id = sheet.cell(row=i,column=1).value
            case.title = sheet.cell(row=i, column=2).value
            case.url = sheet.cell(row=i, column=3).value
            case.data = sheet.cell(row=i, column=4).value
            case.method = sheet.cell(row=i, column=5).value
            case.expected = sheet.cell(row=i, column=6).value
            cases.append(case)
        return cases

    # 获取sheet_name
    def get_sheet_names(self):
        return self.wb.sheetnames

    def write_by_case_id(self,sheet_name,case_id,actual,result):    # case_id为行，column为回写列，value为回写的值
        sheet = self.wb[sheet_name]
        max_row = sheet.max_row
        for r in range(2, max_row+1):
            case_id_r = sheet.cell(r, 1).value  # 取r行第1列获取case_id列
            if case_id_r == case_id:   #判断excel里面取到的当前行的case_id是否等于传入的case_id值
                sheet.cell(r, 7).value = actual
                sheet.cell(r, 8).value = result
                self.wb.save(filename=self.file_name)
                break




if __name__ == '__main__':
    print('comming')
    #  测试一下DoExcel类
    do_excel = DoExcel(contants.data_dir)  # 实例化一个DoExcel对象
    sheet_names = do_excel.get_sheet_names()  # 获取到work boot里面所有的sheet名称的列表
    print("sheet 名称列表：", sheet_names)
    case_list = ['login', 'register']  # 定义一个执行测试用例的列表
    for sheet_name in sheet_names:
        if sheet_name in case_list:  # 如果当前的这个sheet_name 不在可执行的case_list里面，就不执行
            cases = do_excel.get_case(sheet_name)
            print(sheet_name + ' 测试用例个数：', len(cases))
            for case in cases:  # 遍历测试用例列表，每进for一次，就取一个case实例
                print("case信息：", case.__dict__)  # 打印case信息
                data = eval(case.data)  # Excel里面取到data是一个字符串，使用eval函数将字符串转换成字典
                url = ReadConfig().get('test_api', 'url_pre') + case.url
                resp = HttpRequest(method=case.method, url=url, data=data)  # 通过封装的Request类来完成接口的调用
                print('status_code:', resp.get_status_code())  # 打印响应码
                resp_dict = resp.get_json()  # 获取请求响应，字典
                # 通过json.dumps函数将字典转换成格式化后的字符串
                resp_text = json.dumps(resp_dict, ensure_ascii=False, indent=4)
                print('response: ', resp_text)  # 打印响应
                # 判断接口响应是否和Excel里面expected的值是否一致
                if case.expected == resp.get_text():
                    print("result : PASS")
                    do_excel.write_by_case_id(sheet_name=sheet_name, case_id=case.case_id, actual=resp.get_text(),
                                                   result='PASS')  # 期望结果与实际结果一致，就写入PASS到result这个单元格
                else:
                    print("result : FAIL")
                    do_excel.write_by_case_id(sheet_name=sheet_name, case_id=case.case_id,
                                                   actual=resp.get_text(),
                                                   result='FAIL')  # 期望结果与实际结果一致，就写入FAIL到result这个单元格






























